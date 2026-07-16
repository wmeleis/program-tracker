#!/usr/bin/env python3
"""Generate a trustworthy Banner↔CIM code-discrepancy workbook from the
AUTHORITATIVE reconciliation (`banner_reconciliation.code_mismatch` in
data/reports/portfolio_mismatches.json), which already:
  • excludes CPS "P-" quarter programs (going away) on both sides, and
  • shows the ACTUAL current Banner code(s) for the program (matched by
    subject+degree when the CIM code isn't found verbatim).

This replaces the old hand-made ~/Downloads/banner_cim_discrepancies.xlsx, whose
"Banner code" column was mis-derived (most of its codes didn't exist in Banner).

Writes data/reports/banner_cim_discrepancies.xlsx.  Run: python3 banner_cim_report.py
"""
import os
import json
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill

_DIR = os.path.dirname(os.path.abspath(__file__))
_MISMATCHES = os.path.join(_DIR, 'data', 'portfolio_mismatches.json')
_DB = os.path.join(_DIR, 'data', 'tracker.db')
_OUT = os.path.join(_DIR, 'data', 'reports', 'banner_cim_discrepancies.xlsx')


def _credential(conn, cim_code, program):
    """Best-effort credential label from the CIM program's degree, else the
    program name's trailing degree token."""
    row = conn.execute("SELECT degree FROM programs WHERE banner_code=? AND degree!='' LIMIT 1",
                       (cim_code,)).fetchone()
    if row and row[0]:
        return row[0]
    tail = (program or '').rsplit(',', 1)[-1].strip()
    return tail


_HFILL = PatternFill('solid', fgColor='1F3864')


def _style_header(ws):
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = _HFILL
    ws.freeze_panes = 'A2'


def generate():
    with open(_MISMATCHES) as f:
        br = json.load(f).get('banner_reconciliation', {})
    cm = br.get('code_mismatch', []) or []
    cd = br.get('campus_diff', []) or []
    conn = sqlite3.connect(_DB)
    wb = openpyxl.Workbook()

    # Sheet 1 — code mismatch
    ws = wb.active
    ws.title = 'Banner-CIM code mismatch'
    ws.append(['Program', 'Credential', 'CIM Code', 'Banner Code(s)'])
    _style_header(ws)
    for e in sorted(cm, key=lambda x: x.get('program', '')):
        prog = e.get('program', '')
        ws.append([prog, _credential(conn, e.get('cim_code', ''), prog),
                   e.get('cim_code', ''), e.get('banner_code', '')])
    for i, w in enumerate([46, 16, 16, 60], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Sheet 2 — campus differences (per program: where Banner and CIM disagree)
    ws2 = wb.create_sheet('Campus differences')
    ws2.append(['Program (all campuses)', 'Only in Banner', 'Only in CIM',
                'Banner campuses', 'CIM campuses'])
    _style_header(ws2)
    for e in sorted(cd, key=lambda x: x.get('program', '')):
        ws2.append([e.get('program', ''),
                    ', '.join(e.get('only_banner', [])) or '—',
                    ', '.join(e.get('only_portfolio', [])) or '—',
                    ', '.join(e.get('banner_campuses', [])),
                    ', '.join(e.get('cim_campuses', []))])
    for i, w in enumerate([46, 22, 22, 42, 42], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(_OUT)
    conn.close()
    print(f"Wrote {len(cm)} code-mismatch + {len(cd)} campus-difference rows to {_OUT}")
    return len(cm), len(cd)


if __name__ == '__main__':
    generate()
