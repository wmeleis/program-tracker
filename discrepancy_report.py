#!/usr/bin/env python3
"""Consolidated discrepancy report for the Program Portfolio.

Gathers every discrepancy type we track — SVT↔CIM mapping issues, Banner↔portfolio
reconciliation, concentration→college mismatches, and UIP correlation — into one
place. Two consumers share the same gathering logic (`gather_discrepancies`):

  • the weekly xlsx report (`write_xlsx`, written to data/reports/), and
  • the in-app Discrepancies view (Flask /api/discrepancies).

"New since last report" is computed against a stored snapshot
(data/reports/discrepancy_report_state.json). Only GENERATING a report advances
that snapshot; viewing in-app is read-only, so the in-app "new" set always matches
the most recently generated report.

CLI:
  python3 discrepancy_report.py            # generate the xlsx + advance state
  python3 discrepancy_report.py --notify   # + macOS notification if new appeared
  python3 discrepancy_report.py --no-uip   # skip UIP (no Chrome/SharePoint needed)
"""
import os
import re
import json
import time
import argparse
from datetime import datetime

_DIR = os.path.dirname(os.path.abspath(__file__))
_MISMATCHES = os.path.join(_DIR, 'data', 'portfolio_mismatches.json')
_REPORT_DIR = os.path.join(_DIR, 'data', 'reports')
_STATE_PATH = os.path.join(_REPORT_DIR, 'discrepancy_report_state.json')
_DB_PATH = os.path.join(_DIR, 'data', 'tracker.db')


# ---------------------------------------------------------------------------
# Gather
# ---------------------------------------------------------------------------

def _load_mismatches():
    try:
        with open(_MISMATCHES) as f:
            return json.load(f)
    except Exception:
        return {}


def _svt_changed_rows():
    """Entries whose MAPPING fields drifted since a prior review (change_detail
    present and still unreviewed). Excludes the day-one unreviewed backlog — that
    belongs in the Mappings editor, not a discrepancy report."""
    try:
        import database
        seen = database.get_all_svt_seen()
    except Exception:
        return []
    out = []
    for k, v in seen.items():
        last_changed = v.get('last_changed', '') or ''
        last_reviewed = v.get('last_reviewed', '') or ''
        flagged = (not last_reviewed) or (last_changed > last_reviewed)
        try:
            diff = json.loads(v.get('change_detail_json') or '[]')
        except Exception:
            diff = []
        if flagged and diff:
            out.append({
                'svt_key': k,
                'changed_fields': ', '.join(d['field'] for d in diff),
                'detail': '; '.join(f"{d['field']}: \"{d['old']}\" → \"{d['new']}\"" for d in diff),
                'last_changed': last_changed[:10],
            })
    return sorted(out, key=lambda r: r['svt_key'])


def _uip_rows(skip_uip):
    """UIP correlation discrepancies (best-effort). Uses the last-downloaded UIP
    workbook if present; only attempts a fresh download when it's missing."""
    if skip_uip:
        return [{'status': 'skipped', 'discrepancy': 'UIP correlation skipped (--no-uip).'}]
    try:
        import uip_correlate
    except Exception as e:
        return [{'status': 'error', 'discrepancy': f'UIP module unavailable: {e}'}]
    xlsx = os.path.join(_DIR, 'data', 'portfolio_feeds', 'uip_program_information.xlsx')
    try:
        if not os.path.exists(xlsx):
            xlsx = uip_correlate.download_uip_xlsx(xlsx)
        if not xlsx or not os.path.exists(xlsx):
            return [{'status': 'unavailable',
                     'discrepancy': 'UIP workbook could not be downloaded (Chrome/SharePoint session?).'}]
        return [{'discrepancy': s} for s in uip_correlate.correlate(xlsx, _DB_PATH)]
    except Exception as e:
        return [{'status': 'error', 'discrepancy': f'UIP correlation failed: {e}'}]


def gather_discrepancies(skip_uip=False):
    """Return an ordered list of sections. Each section:
       {key, title, columns:[(header, field)], rows:[dict], id_fn}."""
    mm = _load_mismatches()
    br = mm.get('banner_reconciliation', {}) or {}

    def _svt_id(r):
        return r.get('svt_key') or f"{r.get('original_name', r.get('source_name', ''))}|{r.get('campus', r.get('source_campus', ''))}"

    sections = [
        dict(key='svt_added', title='SVT — added (no CIM match)',
             columns=[('SVT Name', 'original_name'), ('Synthesized as', 'cim_format'),
                      ('Campus', 'campus'), ('Intake ID', 'svt_key')],
             rows=mm.get('svt_added', []), id_fn=_svt_id),
        dict(key='svt_pending', title='SVT — pending analysis',
             columns=[('SVT Name', 'source_name'), ('Campus', 'campus'),
                      ('Reason', 'reason'), ('Intake ID', 'svt_key')],
             rows=mm.get('svt_pending_analysis', []), id_fn=_svt_id),
        dict(key='svt_mismatch', title='SVT — mismatch (needs coordination)',
             columns=[('SVT Name', 'source_name'), ('Code', 'source_code'),
                      ('Campus', 'source_campus'), ('Best guess', 'best_guess'), ('Intake ID', 'svt_key')],
             rows=mm.get('svt_mismatches', []), id_fn=_svt_id),
        dict(key='svt_changed', title='SVT — mapping changed since review',
             columns=[('Intake ID', 'svt_key'), ('Changed fields', 'changed_fields'),
                      ('Detail', 'detail'), ('Changed on', 'last_changed')],
             rows=_svt_changed_rows(), id_fn=lambda r: r['svt_key']),
        dict(key='banner_missing_portfolio', title='Banner — in Banner, missing from portfolio',
             columns=[('Banner Code', 'banner_code'), ('Name', 'name')],
             rows=br.get('missing_in_portfolio', []), id_fn=lambda r: r.get('banner_code') or r.get('name')),
        dict(key='banner_missing_banner', title='Banner — in portfolio, missing from Banner',
             columns=[('Program', 'program'), ('Banner Code', 'banner_code')],
             rows=br.get('missing_in_banner', []), id_fn=lambda r: r.get('banner_code') or r.get('program')),
        dict(key='banner_code_mismatch', title='Banner — code mismatch',
             columns=[('Program', 'program'), ('CIM Code', 'cim_code'), ('Banner Code', 'banner_code')],
             rows=br.get('code_mismatch', []), id_fn=lambda r: r.get('program')),
        dict(key='banner_campus_diff', title='Banner — campus differences',
             columns=[('Program', 'program'), ('Banner Code', 'banner_code'),
                      ('Only in portfolio', 'only_portfolio'), ('Only in Banner', 'only_banner')],
             rows=br.get('campus_diff', []), id_fn=lambda r: r.get('program')),
        dict(key='conc_college', title='Concentration → college (CIM vs Banner)',
             columns=[('Program', 'program'), ('Banner Code', 'banner_code'),
                      ('CIM-only', 'cim_only'), ('Banner-only', 'banner_only')],
             rows=mm.get('concentration_college_discrepancies', []), id_fn=lambda r: r.get('program')),
        dict(key='uip', title='UIP correlation',
             columns=[('Discrepancy', 'discrepancy')],
             rows=_uip_rows(skip_uip), id_fn=lambda r: r.get('discrepancy', '')),
    ]
    return sections


# ---------------------------------------------------------------------------
# Deltas
# ---------------------------------------------------------------------------

def _load_state():
    try:
        with open(_STATE_PATH) as f:
            return json.load(f)
    except Exception:
        return {}


def annotate_deltas(sections, prev_state):
    """Mark each row `_new` when its id wasn't in the previous report's snapshot,
    and compute per-section counts. `prev_state` is {section_key: [ids]}. Returns
    (summary, new_state) where new_state can be persisted after generation."""
    prev = (prev_state or {}).get('sections', {})
    first_run = not prev
    summary = []
    new_state = {}
    for sec in sections:
        prev_ids = set(prev.get(sec['key'], []))
        cur_ids = []
        new_count = 0
        for row in sec['rows']:
            rid = str(sec['id_fn'](row))
            cur_ids.append(rid)
            # On the very first run nothing is "new" (there's no prior baseline).
            is_new = (not first_run) and (rid not in prev_ids)
            row['_new'] = is_new
            if is_new:
                new_count += 1
        resolved = len(prev_ids - set(cur_ids)) if not first_run else 0
        summary.append({'key': sec['key'], 'title': sec['title'],
                        'count': len(sec['rows']), 'new': new_count, 'resolved': resolved})
        new_state[sec['key']] = cur_ids
    return summary, {'generated_at': datetime.now().isoformat(), 'sections': new_state}


# ---------------------------------------------------------------------------
# xlsx
# ---------------------------------------------------------------------------

def _cellval(v):
    if isinstance(v, list):
        return ', '.join(str(x) for x in v)
    return '' if v is None else str(v)


def write_xlsx(sections, summary, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill('solid', fgColor='1E40AF')
    hdr_font = Font(color='FFFFFF', bold=True)
    new_fill = PatternFill('solid', fgColor='FEF3C7')
    bold = Font(bold=True)

    # Summary sheet
    ws = wb.active
    ws.title = 'Summary'
    ws.append(['Discrepancy type', 'Count', 'New since last', 'Resolved since last'])
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font
    total = 0
    for s in summary:
        ws.append([s['title'], s['count'], s['new'], s['resolved']])
        total += s['count']
        if s['new']:
            ws.cell(row=ws.max_row, column=3).fill = new_fill
    ws.append([])
    ws.append(['TOTAL', total, sum(s['new'] for s in summary), sum(s['resolved'] for s in summary)])
    for c in ws[ws.max_row]:
        c.font = bold
    ws.append([])
    ws.append([f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} — amber = new since the previous report."])
    ws.column_dimensions['A'].width = 46
    for col in ('B', 'C', 'D'):
        ws.column_dimensions[col].width = 18
    ws.freeze_panes = 'A2'

    # One sheet per section
    used = {'Summary'}
    for sec in sections:
        title = re.sub(r'[\\/*?:\[\]]', ' ', sec['title'])[:28].strip()
        base = title or sec['key']
        name, i = base, 2
        while name in used:
            name = f"{base[:26]} {i}"; i += 1
        used.add(name)
        ws = wb.create_sheet(name)
        headers = ['New'] + [h for h, _ in sec['columns']]
        ws.append(headers)
        for c in ws[1]:
            c.fill = hdr_fill; c.font = hdr_font
        for row in sec['rows']:
            vals = ['NEW' if row.get('_new') else ''] + [_cellval(row.get(f)) for _, f in sec['columns']]
            ws.append(vals)
            if row.get('_new'):
                for c in ws[ws.max_row]:
                    c.fill = new_fill
        ws.column_dimensions['A'].width = 6
        for idx, (h, _) in enumerate(sec['columns'], start=2):
            letter = openpyxl.utils.get_column_letter(idx)
            ws.column_dimensions[letter].width = min(60, max(16, len(h) + 4))
        ws.freeze_panes = 'A2'
        for c in ws[1]:
            c.alignment = Alignment(vertical='top')

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def generate(skip_uip=False, advance_state=True):
    """Full run: gather → delta vs stored state → write dated + latest xlsx →
    (optionally) advance the state snapshot. Returns a dict with paths + summary."""
    os.makedirs(_REPORT_DIR, exist_ok=True)
    sections = gather_discrepancies(skip_uip=skip_uip)
    summary, new_state = annotate_deltas(sections, _load_state())
    dated = os.path.join(_REPORT_DIR, f"discrepancy_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    latest = os.path.join(_REPORT_DIR, 'discrepancy_report_latest.xlsx')
    write_xlsx(sections, summary, dated)
    write_xlsx(sections, summary, latest)
    if advance_state:
        with open(_STATE_PATH, 'w') as f:
            json.dump(new_state, f, indent=1)
    return {'dated_path': dated, 'latest_path': latest, 'summary': summary,
            'total': sum(s['count'] for s in summary),
            'total_new': sum(s['new'] for s in summary)}


def view_data(skip_uip=True):
    """Read-only gather + delta for the in-app view. Does NOT advance the state
    snapshot, so 'new since last report' matches the last generated xlsx. UIP is
    skipped by default (it needs Chrome/SharePoint) — generate the report for UIP."""
    sections = gather_discrepancies(skip_uip=skip_uip)
    summary, _ = annotate_deltas(sections, _load_state())
    return {'sections': [{'key': s['key'], 'title': s['title'],
                          'columns': [h for h, _ in s['columns']],
                          'fields': [f for _, f in s['columns']],
                          'rows': [{**{f: _cellval(r.get(f)) for _, f in s['columns']},
                                    '_new': bool(r.get('_new'))} for r in s['rows']]}
                         for s in sections],
            'summary': summary,
            'state_generated_at': (_load_state() or {}).get('generated_at', '')}


def _notify(title, msg):
    try:
        os.system(f'''osascript -e 'display notification "{msg}" with title "{title}"' ''')
    except Exception:
        pass


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--notify', action='store_true', help='macOS notification if new discrepancies appeared')
    ap.add_argument('--no-uip', action='store_true', help='skip UIP correlation (no Chrome/SharePoint)')
    args = ap.parse_args()
    res = generate(skip_uip=args.no_uip)
    print(f"Discrepancy report: {res['total']} total, {res['total_new']} new since last")
    for s in res['summary']:
        flag = f"  (+{s['new']} new, -{s['resolved']} resolved)" if (s['new'] or s['resolved']) else ''
        print(f"  {s['title']}: {s['count']}{flag}")
    print(f"Written: {res['latest_path']}")
    if args.notify and res['total_new']:
        _notify('Portfolio discrepancies', f"{res['total_new']} new since last report ({res['total']} total)")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
