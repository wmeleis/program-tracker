#!/usr/bin/env python3
"""UIP roster ↔ CIM correlation — weekly data-quality diff.

Downloads the UIP Program Information workbook from SharePoint (via the logged-in
Chrome session, same bridge the regulatory workbooks use), then compares each UIP
interdisciplinary program against the CIM tracker on three dimensions:

  1. Concentrations — UIP track list vs CIM curriculum concentrations (Boston).
  2. Campuses       — UIP deployment campuses vs CIM deployment campuses (exact).
  3. Launch terms   — only meaningful for NEW programs; for Change/Inactivation
                      proposals the CIM eff_term is the change term, not the
                      original launch, so those are treated as benign.

Also surfaces CIM inactivations not reflected on the still-active UIP roster.

Emits ONLY discrepancies. Writes a markdown report to --out (default
data/uip_discrepancies.md) and prints it. Exit code 0 always (so the scheduler
can decide what to send); the report starts with "NO DISCREPANCIES" when clean.

Usage: python3 uip_correlate.py [--out PATH] [--xlsx PATH]
Run from the project directory (needs scraper.py, tracker.db, openpyxl).
"""
import argparse, base64, json, os, re, sqlite3, sys, time
from datetime import datetime

_UIP_SITE = "https://northeastern.sharepoint.com/sites/UIPProgramCommittees"
_UIP_TAB_MATCH = "sharepoint.com/sites/UIPProgramCommittees"
_UIP_SERVER_REL = "/sites/UIPProgramCommittees/Shared Documents/General/UIP Program Information.xlsx"

_SUF = {'10':'Fall','12':'Fall','14':'Fall','15':'Fall','18':'Fall',
        '25':'Winter','28':'Winter',
        '30':'Spring','32':'Spring','34':'Spring','35':'Spring','38':'Spring',
        '40':'Summer','50':'Summer','52':'Summer','54':'Summer','55':'Summer','58':'Summer','60':'Summer'}


def _decode_term(code):
    c = str(code or '')
    if len(c) != 6 or not c.isdigit() or c[4:] not in _SUF:
        return ''
    s = _SUF[c[4:]]; lead = int(c[:4])
    return f"{s} {lead - 1 if s == 'Fall' else lead}"


def _nname(s):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', ' ', (s or '').lower().split('(')[0])).strip()


def _nconc(s):
    s = (s or '').lower()
    s = re.split(r'[—(]', s)[0]
    s = re.sub(r'\bconcentration\b', ' ', s)
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', ' ', s)).strip()


def download_uip_xlsx(dest):
    """Fetch the workbook via Chrome/SharePoint REST. Returns dest path or None."""
    from scraper import run_js_in_tab
    # Ensure a tab on the UIP site exists (open one if not) so the fetch is same-origin + authenticated.
    href = run_js_in_tab(_UIP_TAB_MATCH, "location.href", match_by='url', timeout=10)
    if not href or 'UIPProgramCommittees' not in str(href):
        os.system('osascript -e \'tell application "Google Chrome"\n'
                  'if (count of windows) = 0 then make new window\n'
                  f'tell window 1 to make new tab with properties {{URL:"{_UIP_SITE}"}}\n'
                  'end tell\' >/dev/null 2>&1')
        for _ in range(20):
            time.sleep(2)
            href = run_js_in_tab(_UIP_TAB_MATCH, "location.href", match_by='url', timeout=10)
            if href and 'UIPProgramCommittees' in str(href):
                break
        else:
            print("  UIP SharePoint tab unavailable / not authenticated — skipping")
            return None
    kickoff = '''
(function(){
  window.__uip=null; window.__uip_status="running";
  var seg="%s".split('/').map(encodeURIComponent).join('/');
  var url=location.origin+"/sites/UIPProgramCommittees/_api/web/GetFileByServerRelativeUrl('"+seg+"')/$value";
  var xhr=new XMLHttpRequest(); xhr.open("GET",url,true); xhr.responseType="arraybuffer";
  xhr.onload=function(){
    if(xhr.status>=200&&xhr.status<300){
      var b=new Uint8Array(xhr.response),bin="",step=32768;
      for(var i=0;i<b.length;i+=step){bin+=String.fromCharCode.apply(null,b.subarray(i,i+step));}
      window.__uip={status:xhr.status,len:b.length,b64:btoa(bin)}; window.__uip_status="done";
    } else { window.__uip={status:xhr.status,error:"http"}; window.__uip_status="done"; }
  };
  xhr.onerror=function(){window.__uip={error:"network"};window.__uip_status="done";};
  xhr.send(); return "fired";
})();
''' % _UIP_SERVER_REL
    if run_js_in_tab(_UIP_TAB_MATCH, kickoff, match_by='url', timeout=30) in (None, 'missing value'):
        print("  could not start UIP download")
        return None
    for _ in range(30):
        time.sleep(2)
        if run_js_in_tab(_UIP_TAB_MATCH, 'window.__uip_status||"?"', match_by='url', timeout=10) == "done":
            break
    meta = run_js_in_tab(_UIP_TAB_MATCH,
                         'JSON.stringify(window.__uip?{status:window.__uip.status,b64len:(window.__uip.b64||"").length,error:window.__uip.error||null}:null)',
                         match_by='url', timeout=15)
    try:
        m = json.loads(meta) if meta and meta != 'null' else None
    except Exception:
        m = None
    if not m or m.get('error') or not m.get('b64len'):
        print(f"  UIP download failed ({m})")
        return None
    parts = []
    for off in range(0, m['b64len'], 200000):
        parts.append(run_js_in_tab(_UIP_TAB_MATCH, f'window.__uip.b64.substr({off},200000)', match_by='url', timeout=30))
    data = base64.b64decode(''.join(p for p in parts if p and p != 'missing value'))
    with open(dest, 'wb') as f:
        f.write(data)
    return dest


def correlate(xlsx_path, db_path='data/tracker.db'):
    """Return a list of discrepancy strings."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    conn = sqlite3.connect(db_path); conn.row_factory = sqlite3.Row
    pf = conn.execute("SELECT pp.*, p.banner_code, p.eff_term FROM portfolio_programs pp "
                      "JOIN programs p ON p.id = pp.cim_program_id").fetchall()
    dep_campus, boston = {}, {}
    for r in pf:
        key = _nname(r['program_name'])
        dep_campus.setdefault(key, set()).add(r['campus'] or 'Boston')
        if r['banner_code']:
            dep_campus.setdefault(r['banner_code'].upper(), set()).add(r['campus'] or 'Boston')
        if r['campus'] == 'Boston' or '(' not in (r['program_name'] or ''):
            boston.setdefault(key, r)
            if r['banner_code']:
                boston.setdefault(r['banner_code'].upper(), r)
    # inactivation deployments per program key/code
    inactivating = {}
    for r in pf:
        if r['cim_change_type'] == 'Inactivation':
            inactivating.setdefault(_nname(r['program_name']), []).append(r['campus'] or 'Boston')
            if r['banner_code']:
                inactivating.setdefault(r['banner_code'].upper(), []).append(r['campus'] or 'Boston')

    TERM = re.compile(r'^(Fall|Spring|Summer|Winter)\s+\d{4}$', re.I)
    disc = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        body = [r for r in list(ws.iter_rows(values_only=True))
                if r and r[1] and str(r[1]).strip() not in ('Tracks/Concentrations', '')]
        code = ''
        for r in body:
            if r[0]:
                m = re.match(r'\s*([A-Z]{2,6}-[A-Z0-9]{2,6})\s*:', str(r[0]))
                if m:
                    code = m.group(1).upper(); break
        cim = boston.get(code) or boston.get(_nname(sn))
        if not cim:
            disc.append(f"**{sn}**: not found in CIM (no banner-code/name match).")
            continue
        key = code if code in dep_campus else _nname(sn)
        # 1. concentrations (Boston vs Boston)
        bos_tracks = [str(r[1]).strip() for r in body
                      if not (len(r) > 3 and r[3]) or str(r[3]).strip() == 'Boston']
        try:
            cimc = [c['name'] for c in json.loads(cim['concentrations_json'] or '[]')]
        except Exception:
            cimc = []
        if cimc:
            uset = {_nconc(x) for x in bos_tracks}
            cset = {_nconc(x) for x in cimc}
            uonly = [x for x in bos_tracks if _nconc(x) not in cset]
            conly = [x for x in cimc if _nconc(x) not in uset]
            if uonly or conly:
                msg = f"**{sn}** concentrations differ:"
                if uonly: msg += f" UIP-only={uonly};"
                if conly: msg += f" CIM-only={conly};"
                disc.append(msg)
        # 2. campuses (exact)
        ucamp = {str(r[3]).strip() for r in body if len(r) > 3 and r[3] and str(r[3]).strip() != 'Campus'}
        ccamp = set(dep_campus.get(key, set()))
        if ucamp and ucamp != ccamp:
            disc.append(f"**{sn}** campuses differ: UIP-only={sorted(ucamp - ccamp)}, "
                        f"CIM-only={sorted(ccamp - ucamp)}.")
        # 3. launch term — only when the CIM proposal is a NEW program
        if (cim['cim_change_type'] or '') in ('New', 'Added'):
            uterms = {str(r[6]).strip() for r in body if len(r) > 6 and TERM.match(str(r[6]).strip() or '')}
            cimt = _decode_term(cim['eff_term'])
            if uterms and cimt and cimt not in uterms:
                disc.append(f"**{sn}** launch term: CIM (new program) eff_term {cimt} "
                            f"not among UIP intake terms {sorted(uterms)}.")
        # 4. CIM inactivation not reflected on the (active) UIP roster
        inact = sorted(set(inactivating.get(key, [])))
        if inact:
            disc.append(f"**{sn}**: CIM has an INACTIVATION proposal for campus(es) {inact}, "
                        f"but the program is still on the active UIP roster.")
    return disc


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', default='data/uip_discrepancies.md')
    ap.add_argument('--xlsx', default=None, help='use a local .xlsx instead of downloading')
    args = ap.parse_args()
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    xlsx = args.xlsx
    if not xlsx:
        xlsx = download_uip_xlsx(os.path.join('data', 'portfolio_feeds', 'uip_program_information.xlsx'))
        if not xlsx:
            print("Could not obtain the UIP workbook; leaving previous report untouched.")
            return 0
    disc = correlate(xlsx)
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    if disc:
        report = f"# UIP ↔ CIM discrepancies — {stamp}\n\n" + '\n'.join(f"- {d}" for d in disc) + "\n"
    else:
        report = f"# UIP ↔ CIM discrepancies — {stamp}\n\nNO DISCREPANCIES — UIP roster and CIM agree on concentrations, campuses, and launch terms.\n"
    with open(args.out, 'w') as f:
        f.write(report)
    print(report)
    return 0


if __name__ == '__main__':
    sys.exit(main())
