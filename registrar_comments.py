#!/usr/bin/env python3
"""Extract reviewer comments from graduate programs' CIM workflow pages.

Sweeps every graduate program that went through the workflow (active + completed),
pulls the `wfcomments-cmt` reviewer-comment stream from each `/programadmin/{id}/`
page, filters to a rolling window (default 2 years), and writes a raw corpus:

  data/reports/registrar_comments_corpus.json   — structured, for the guide steps
  data/reports/registrar_comments.xlsx          — By-commenter summary + all comments

Comments are stamped by CourseLeaf with author, username, and a GMT timestamp:
  <div class="wfcomments-cmt"><b>Heather Daly (h.daly) (<span class="timestamp">
  Wed, 09 Jul 2026 14:02:11 GMT</span>):</b> Registrar's Office note: …</div>
"""
import os
import re
import sys
import json
import html
import sqlite3
import argparse
from datetime import datetime, timedelta, timezone
from email.utils import parsedate_to_datetime

import cim_http

_DIR = os.path.dirname(os.path.abspath(__file__))
_DB = os.path.join(_DIR, 'data', 'tracker.db')
_REPORT_DIR = os.path.join(_DIR, 'data', 'reports')

# Registrar's Office reviewer usernames (confirmed by Waleed 2026-07-12). The
# corpus keeps ALL commenters, but this set flags the Registrar's Office voice —
# the authoritative source for the compliance guide. Explicitly NOT Registrar:
# v.wallace (BCHS), k.mellor (college), w.meleis / program directors / deans.
REGISTRAR_USERS = {
    'h.daly', 'da.rogers', 'm.depaula', 'm.boudreault', 'ben.joseph',
    'c.cornwall', 'm.standig', 'm.couch-hrinda', 'm.daigle',
}

_CMT_RE = re.compile(
    r'<div class="wfcomments-cmt"><b>(?P<author>.*?)\((?P<user>[\w.\-]+)\)\s*'
    r'\(<span class="timestamp">(?P<ts>.*?)</span>\):</b>(?P<text>.*?)</div>',
    re.S)


def _clean(t):
    return html.unescape(re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', '', t))).strip()


def _parse_ts(ts):
    try:
        dt = parsedate_to_datetime(ts.strip())
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


def graduate_workflow_programs():
    c = sqlite3.connect(_DB); c.row_factory = sqlite3.Row
    rows = c.execute(
        "SELECT id, name, college, current_step, completion_date FROM programs "
        "WHERE program_type='Graduate' AND (current_step!='' OR completion_date!='') "
        "ORDER BY id").fetchall()
    return [dict(r) for r in rows]


def extract(window_days=730, limit=None, progress=True):
    cutoff = datetime.now(timezone.utc) - timedelta(days=window_days)
    progs = graduate_workflow_programs()
    if limit:
        progs = progs[:limit]
    sess = cim_http.CIMSession()
    comments = []
    fetched = with_cmts = 0
    aborted = False
    for i, p in enumerate(progs):
        html_body = sess.get(f"/programadmin/{p['id']}/")
        if sess.logged_out:
            print("!! CIM session expired mid-sweep — stopping. Re-login and rerun.", file=sys.stderr)
            aborted = True
            break
        if not html_body:
            continue
        fetched += 1
        found = False
        for m in _CMT_RE.finditer(html_body):
            dt = _parse_ts(m.group('ts'))
            if dt is None or dt < cutoff:
                continue
            found = True
            text = _clean(m.group('text'))
            if not text:
                continue
            comments.append({
                'program_id': p['id'],
                'program': p['name'],
                'college': p['college'] or '',
                'current_step': p['current_step'] or '',
                'completed': bool(p['completion_date']),
                'author': _clean(m.group('author')),
                'username': m.group('user'),
                'registrar': m.group('user') in REGISTRAR_USERS,
                'date': dt.strftime('%Y-%m-%d'),
                'datetime': dt.isoformat(),
                'is_rollback': text.lower().startswith('rollback'),
                'comment': text,
            })
        if found:
            with_cmts += 1
        if progress and (i + 1) % 100 == 0:
            print(f"  …{i+1}/{len(progs)} programs, {len(comments)} comments so far", file=sys.stderr)
    comments.sort(key=lambda c: (c['username'].lower(), c['program'].lower(), c['datetime']))
    return {
        'generated_at': datetime.now().isoformat(),
        'window_days': window_days,
        'cutoff': cutoff.strftime('%Y-%m-%d'),
        'programs_swept': len(progs),
        'programs_fetched': fetched,
        'programs_with_comments': with_cmts,
        'total_comments': len(comments),
        'aborted': aborted,
        'comments': comments,
    }


def write_corpus(data):
    os.makedirs(_REPORT_DIR, exist_ok=True)
    jpath = os.path.join(_REPORT_DIR, 'registrar_comments_corpus.json')
    with open(jpath, 'w') as f:
        json.dump(data, f, indent=1)

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from collections import Counter
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill('solid', fgColor='1E40AF'); hdr_font = Font(color='FFFFFF', bold=True)

    ws = wb.active; ws.title = 'By commenter'
    ws.append(['Commenter', 'Username', 'Registrar', 'Comments', 'Substantive', 'Rollbacks/procedural', 'Programs'])
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font
    by_user = {}
    for cm in data['comments']:
        u = by_user.setdefault(cm['username'], {'author': cm['author'], 'reg': cm.get('registrar'), 'n': 0, 'rb': 0, 'progs': set()})
        u['n'] += 1; u['rb'] += 1 if cm['is_rollback'] else 0; u['progs'].add(cm['program_id'])
    for user, u in sorted(by_user.items(), key=lambda kv: -kv[1]['n']):
        ws.append([u['author'], user, 'Y' if u['reg'] else '', u['n'], u['n'] - u['rb'], u['rb'], len(u['progs'])])
    ws.column_dimensions['A'].width = 26; ws.column_dimensions['B'].width = 16
    for col in ('C', 'D', 'E', 'F', 'G'):
        ws.column_dimensions[col].width = 13
    ws.freeze_panes = 'A2'

    ws2 = wb.create_sheet('All comments')
    cols = [('Commenter', 'author'), ('Username', 'username'), ('Registrar', 'registrar'),
            ('Date', 'date'), ('Program', 'program'), ('College', 'college'), ('Prog ID', 'program_id'),
            ('Rollback', 'is_rollback'), ('Comment', 'comment')]
    ws2.append([h for h, _ in cols])
    for c in ws2[1]:
        c.fill = hdr_fill; c.font = hdr_font
    for cm in data['comments']:
        ws2.append([('Y' if cm.get(f) else '') if f in ('is_rollback', 'registrar') else cm[f] for _, f in cols])
    widths = [24, 15, 9, 11, 40, 10, 8, 9, 100]
    for idx, w in enumerate(widths, start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    ws2.freeze_panes = 'A2'
    xpath = os.path.join(_REPORT_DIR, 'registrar_comments.xlsx')
    wb.save(xpath)
    return jpath, xpath


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--years', type=float, default=2.0)
    ap.add_argument('--limit', type=int, default=None, help='cap program count (testing)')
    args = ap.parse_args()
    data = extract(window_days=int(args.years * 365), limit=args.limit)
    # Safety: never overwrite a good corpus with an empty/partial one when the CIM
    # session dropped. Require a healthy fetch (session live, most programs fetched).
    if data['aborted'] or data['programs_fetched'] < 0.5 * max(1, data['programs_swept']):
        print(f"ABORTED: only {data['programs_fetched']}/{data['programs_swept']} programs fetched "
              f"(CIM session expired?). Corpus NOT overwritten — re-login to CIM and rerun.",
              file=sys.stderr)
        return 1
    jpath, xpath = write_corpus(data)
    print(f"Swept {data['programs_swept']} grad programs "
          f"({data['programs_fetched']} fetched, {data['programs_with_comments']} with comments)")
    print(f"{data['total_comments']} comments in the last {args.years} years")
    print(f"Corpus: {jpath}\n        {xpath}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
